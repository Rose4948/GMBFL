import pickle
import os
import sys

import openpyxl as op
K_size=1
proj = sys.argv[1]
seed = int(sys.argv[2])
lr = float(sys.argv[3])
batch_size = int(sys.argv[4])
NNName = sys.argv[5]
EPOCHS=int(sys.argv[6])+1
model=NNName+'-lr'+str(lr)

def write(epoch,datalist):
    bg = op.load_workbook(proj+'.xlsx') 
    sheet = bg["Sheet"+str(epoch)]
    row=sheet.max_row+1
    for col in range(1,len(datalist)+1):
        sheet.cell(row, col,datalist[col-1])  
    bg.save(proj+'.xlsx')  

versionNum = {'Lang': 65, 'Time': 27, 'Chart': 26, 'Math': 106, 'Closure': 133, 'Mockito': 38,
                   'CommonsCli': 24, 'CommonsCodec': 22, 'CommonsCsv': 12, 'CommonsJXPath': 14,
                   'JacksonCore': 13, 'JacksonDatabind': 39, 'JacksonXml': 5, 'Jsoup': 63,'Cli':38}


t = {}
t[0]=[]
t[1]={}
t[2]=[]
t[3]={}

for i in range(0, versionNum[proj]):
    if not os.path.exists(proj + '_%s_%d_%d_%s_%s.pkl'%(NNName, i, seed, lr, batch_size)):
        continue
    p = pickle.load(open(proj + '_%s_%d_%d_%s_%s.pkl'%(NNName, i,seed, lr, batch_size), 'rb'))
    p=p[i]
    Max_expoch=[len(p[0]),len(p[1]),len(p[2]),len(p[3])]
    Max_expoch=[1,1,len(p[2]),EPOCHS]
    for j in [0,2]:
        for x in range(Max_expoch[j]):
            t[j].append(p[j][x])
    for key in p[1]:
        t[1][key] = p[1][key]
    for j in range(Max_expoch[3]):
        if j not in t[3]:
            t[3][j]={}
        for key in p[3][j]:
            t[3][j][key]=p[3][j][key]


open(proj + '_%s_%d_%s_%s.pkl'%(NNName, seed,lr, batch_size), 'wb').write(pickle.dumps(t))
data= pickle.load(open(proj + '.pkl', 'rb'))

ANS=[0,0,0,0,0]
epoch=10
ranks=t[3][epoch]
for key in ranks:
    ranklist = ranks[key]
    for x in data[key]['ans']:
        rank_x = ranklist.index(x)
        if rank_x<5:
            ANS[rank_x]+=1
print("TOP-1,TOP-3,TOP-5:",ANS[0], ANS[0]+ANS[1]+ANS[2],ANS[0]+ANS[1]+ANS[2]+ANS[3]+ANS[4])